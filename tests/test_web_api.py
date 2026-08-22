from __future__ import annotations

from io import BytesIO
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from rasp.web.app import create_app
from rasp.imports.excel import read_import_workbook


FIXTURES = Path(__file__).parent / "fixtures"
XLSX_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


class WebApiTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary_directory = tempfile.TemporaryDirectory()
        self.database_path = Path(self.temporary_directory.name) / "rasp.sqlite3"
        self.client = TestClient(create_app(database_path=self.database_path))

    def tearDown(self) -> None:
        self.client.close()
        self.temporary_directory.cleanup()

    def upload(self, endpoint: str, fixture: str, **kwargs: object):
        content = (FIXTURES / fixture).read_bytes()
        return self.client.post(
            endpoint,
            files={"file": (fixture, content, XLSX_CONTENT_TYPE)},
            **kwargs,
        )

    def test_home_page_is_accessible_and_hardened(self) -> None:
        response = self.client.get("/")
        stylesheet = self.client.get("/static/styles.css")
        script = self.client.get("/static/app.js")
        favicon = self.client.get("/static/favicon.svg")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(stylesheet.status_code, 200)
        self.assertEqual(script.status_code, 200)
        self.assertEqual(favicon.status_code, 200)
        self.assertIn("Импорт исходных данных", response.text)
        self.assertIn('id="readiness-summary"', response.text)
        self.assertIn('id="academic-cycle-count"', response.text)
        self.assertIn('id="cycle-commission-count"', response.text)
        self.assertIn('id="teacher-replacement-count"', response.text)
        self.assertIn('id="solve-button"', response.text)
        self.assertIn('id="schedule-result"', response.text)
        self.assertIn('id="mode-selector"', response.text)
        self.assertIn('id="manual-workspace"', response.text)
        self.assertIn('id="manual-activate-button"', response.text)
        self.assertIn("renderReadiness(payload.readiness)", script.text)
        self.assertIn('fetch("/api/manual-data/activate"', script.text)
        self.assertIn('fetch("/api/solver/runs"', script.text)
        self.assertIn("renderSchedule(payload)", script.text)
        self.assertEqual(response.headers["x-frame-options"], "DENY")
        self.assertEqual(response.headers["x-content-type-options"], "nosniff")
        self.assertIn("default-src 'self'", response.headers["content-security-policy"])

    def test_preview_returns_counts_and_does_not_create_database(self) -> None:
        response = self.upload("/api/imports/preview", "valid-import.xlsx")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["valid"])
        self.assertEqual(payload["counts"]["teachers"], 1)
        self.assertEqual(payload["counts"]["specialties"], 1)
        self.assertEqual(payload["counts"]["curricula"], 1)
        self.assertEqual(payload["counts"]["disciplines"], 1)
        self.assertEqual(payload["counts"]["students"], 1)
        self.assertEqual(payload["counts"]["buildings"], 1)
        self.assertEqual(payload["counts"]["rooms"], 1)
        self.assertEqual(payload["counts"]["academicYears"], 1)
        self.assertEqual(payload["counts"]["academicCycles"], 1)
        self.assertEqual(payload["counts"]["cycleCommissions"], 0)
        self.assertEqual(payload["counts"]["teacherReplacements"], 0)
        self.assertEqual(payload["counts"]["calendarPeriods"], 1)
        self.assertEqual(payload["counts"]["bellSlots"], 1)
        self.assertEqual(payload["counts"]["calendarExceptions"], 1)
        self.assertEqual(payload["counts"]["resourceUnavailability"], 1)
        self.assertEqual(payload["studentChanges"]["created"], 1)
        self.assertEqual(payload["roomDeficits"], [])
        self.assertTrue(payload["readiness"]["isReady"])
        self.assertEqual(payload["readiness"]["issues"], [])
        self.assertTrue(payload["solverProblem"]["isReady"])
        self.assertEqual(payload["solverProblem"]["lessonDemandCount"], 36)
        self.assertEqual(payload["solverProblem"]["eligibleWeekCount"], 9)
        self.assertEqual(payload["solverProblem"]["placementDomainCount"], 1)
        self.assertEqual(payload["solverProblem"]["placementOptionCount"], 50)
        self.assertEqual(payload["samples"]["groups"][0]["groupCode"], "ИС-101")
        self.assertFalse(self.database_path.exists())

    def test_preview_returns_structured_validation_issues(self) -> None:
        response = self.upload("/api/imports/preview", "formula-import.xlsx")

        self.assertEqual(response.status_code, 422)
        payload = response.json()
        self.assertFalse(payload["valid"])
        self.assertEqual(payload["issues"][0]["code"], "formula_forbidden")

    def test_excel_and_manual_routes_reject_same_replacement_scope(self) -> None:
        source = BytesIO((FIXTURES / "valid-import.xlsx").read_bytes())
        workbook = load_workbook(source)
        workbook["Преподаватели"].append(
            ["T-002", "Петров Пётр Петрович", None, "staff", 0, 900, 8, 5, "MAIN", True]
        )
        replacements = workbook.create_sheet("Замены преподавателей")
        replacements.append(
            [
                "Код замены",
                "Учебный год",
                "Код основного преподавателя",
                "Код замещающего преподавателя",
                "Дата начала",
                "Дата окончания",
                "Код строки нагрузки",
                "Причина",
            ]
        )
        replacements.append(
            [
                "REP-001",
                "2026/2027",
                "T-002",
                "T-001",
                "2026-09-01",
                "2026-09-30",
                "W-001",
                None,
            ]
        )
        changed = BytesIO()
        workbook.save(changed)
        workbook.close()
        content = changed.getvalue()
        workbook_path = Path(self.temporary_directory.name) / "invalid-replacement.xlsx"
        workbook_path.write_bytes(content)
        batch = read_import_workbook(workbook_path)

        excel_response = self.client.post(
            "/api/imports/preview",
            files={"file": (workbook_path.name, content, XLSX_CONTENT_TYPE)},
        )
        manual_response = self.client.post(
            "/api/manual-data/activate",
            json={"batch": batch.model_dump(mode="json")},
        )

        expected_code = "replacement_workload_teacher_mismatch"
        self.assertEqual(excel_response.status_code, 422)
        self.assertEqual(manual_response.status_code, 422)
        self.assertIn(
            expected_code,
            {issue["code"] for issue in excel_response.json()["issues"]},
        )
        self.assertIn(
            expected_code,
            {issue["code"] for issue in manual_response.json()["issues"]},
        )

    def test_activate_then_status_reports_active_version(self) -> None:
        activated = self.upload("/api/imports/activate", "valid-import.xlsx")
        status = self.client.get("/api/status")

        self.assertEqual(activated.status_code, 201)
        self.assertEqual(activated.json()["versionId"], 1)
        self.assertEqual(status.status_code, 200)
        self.assertEqual(status.json()["activeVersionId"], 1)
        self.assertEqual(status.json()["counts"]["disciplines"], 1)
        self.assertEqual(status.json()["counts"]["students"], 1)
        self.assertEqual(status.json()["counts"]["rooms"], 1)
        self.assertEqual(status.json()["counts"]["academicYears"], 1)
        self.assertEqual(status.json()["counts"]["cycleCommissions"], 0)
        self.assertEqual(status.json()["counts"]["teacherReplacements"], 0)

        readiness = self.client.get("/api/readiness")
        self.assertEqual(readiness.status_code, 200)
        self.assertTrue(readiness.json()["isReady"])
        self.assertEqual(readiness.json()["errorCount"], 0)

        solver_problem = self.client.get("/api/solver/problem")
        self.assertEqual(solver_problem.status_code, 200)
        self.assertTrue(solver_problem.json()["isReady"])
        self.assertEqual(solver_problem.json()["lessonDemandCount"], 36)
        self.assertEqual(solver_problem.json()["placementOptionCount"], 50)

    def test_readiness_without_active_version_is_conflict(self) -> None:
        response = self.client.get("/api/readiness")

        self.assertEqual(response.status_code, 409)
        self.assertEqual(response.json()["error"]["code"], "no_active_import")

    def test_solver_problem_without_active_version_is_conflict(self) -> None:
        response = self.client.get("/api/solver/problem")

        self.assertEqual(response.status_code, 409)
        self.assertEqual(response.json()["error"]["code"], "no_active_import")

    def test_solver_run_returns_draft_schedule(self) -> None:
        self.upload("/api/imports/activate", "valid-import.xlsx")

        response = self.client.post(
            "/api/solver/runs",
            json={"mode": "complete", "seed": 7, "timeLimitSeconds": 10},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["status"], "feasible")
        self.assertEqual(payload["seed"], 7)
        self.assertEqual(payload["assignmentCount"], 36)
        self.assertEqual(payload["assignments"][0]["demandCode"], "W-001#001")
        self.assertEqual(payload["assignments"][0]["disciplineCode"], "MDK.01.01")
        self.assertTrue(payload["assignments"][0]["occupiedSlotCodes"])

    def test_solver_run_reports_engine_failure_without_private_details(self) -> None:
        self.upload("/api/imports/activate", "valid-import.xlsx")

        with patch(
            "rasp.web.app.CpSatScheduleSolver.solve",
            side_effect=RuntimeError("private runtime detail"),
        ):
            response = self.client.post("/api/solver/runs", json={})

        self.assertEqual(response.status_code, 500)
        payload = response.json()
        self.assertEqual(payload["error"]["code"], "solver_engine_error")
        self.assertIn("RuntimeError", payload["error"]["message"])
        self.assertNotIn("private runtime detail", response.text)

    def test_solver_run_validates_options_and_requires_active_version(self) -> None:
        missing = self.client.post("/api/solver/runs", json={})
        invalid = self.client.post(
            "/api/solver/runs", json={"timeLimitSeconds": 301}
        )
        invalid_seed = self.client.post(
            "/api/solver/runs", json={"seed": 2_147_483_648}
        )

        self.assertEqual(missing.status_code, 409)
        self.assertEqual(missing.json()["error"]["code"], "no_active_import")
        self.assertEqual(invalid.status_code, 422)
        self.assertEqual(invalid_seed.status_code, 422)

    def test_manual_data_activation_creates_a_standard_active_version(self) -> None:
        batch = read_import_workbook(FIXTURES / "valid-import.xlsx")

        response = self.client.post(
            "/api/manual-data/activate",
            json={
                "sourceName": "Ручной ввод: тестовый колледж",
                "batch": batch.model_dump(mode="json"),
            },
        )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.json()["versionId"], 1)
        self.assertEqual(response.json()["sourceName"], "Ручной ввод: тестовый колледж")
        readiness = self.client.get("/api/readiness")
        self.assertTrue(readiness.json()["isReady"])

    def test_invalid_manual_data_does_not_replace_active_version(self) -> None:
        valid_batch = read_import_workbook(FIXTURES / "valid-import.xlsx")
        self.client.post(
            "/api/manual-data/activate",
            json={"batch": valid_batch.model_dump(mode="json")},
        )
        invalid_batch = valid_batch.model_copy(
            update={
                "workloads": (
                    valid_batch.workloads[0].model_copy(
                        update={"discipline_code": "UNKNOWN"}
                    ),
                )
            }
        )

        response = self.client.post(
            "/api/manual-data/activate",
            json={"batch": invalid_batch.model_dump(mode="json")},
        )

        self.assertEqual(response.status_code, 422)
        self.assertEqual(response.json()["issues"][0]["code"], "discipline_not_in_curriculum")
        self.assertEqual(self.client.get("/api/status").json()["activeVersionId"], 1)

    def test_incomplete_structurally_valid_manual_data_becomes_working_version(self) -> None:
        valid_batch = read_import_workbook(FIXTURES / "valid-import.xlsx")
        self.client.post(
            "/api/manual-data/activate",
            json={"batch": valid_batch.model_dump(mode="json")},
        )
        incomplete_batch = valid_batch.model_copy(
            update={
                "teachers": (),
                "workloads": (),
                "resource_unavailability": (),
            }
        )

        response = self.client.post(
            "/api/manual-data/activate",
            json={"batch": incomplete_batch.model_dump(mode="json")},
        )

        self.assertEqual(response.status_code, 201)
        self.assertEqual(response.json()["versionId"], 2)
        self.assertEqual(self.client.get("/api/status").json()["activeVersionId"], 2)
        readiness = self.client.get("/api/readiness").json()
        self.assertFalse(readiness["isReady"])
        self.assertIn(
            "missing_teachers",
            {issue["code"] for issue in readiness["issues"]},
        )

    def test_preview_reports_workload_without_matching_room(self) -> None:
        source = BytesIO((FIXTURES / "valid-import.xlsx").read_bytes())
        workbook = load_workbook(source)
        workbook["Аудитории"]["E2"] = 20
        changed = BytesIO()
        workbook.save(changed)
        workbook.close()

        preview = self.client.post(
            "/api/imports/preview",
            files={"file": ("rooms.xlsx", changed.getvalue(), XLSX_CONTENT_TYPE)},
        )

        self.assertEqual(preview.status_code, 200)
        self.assertEqual(preview.json()["roomDeficits"][0]["workloadRowCode"], "W-001")
        self.assertEqual(preview.json()["roomDeficits"][0]["requiredCapacity"], 25)
        self.assertFalse(preview.json()["readiness"]["isReady"])
        self.assertIn(
            "no_suitable_room",
            {issue["code"] for issue in preview.json()["readiness"]["issues"]},
        )

        activated = self.client.post(
            "/api/imports/activate",
            files={"file": ("rooms.xlsx", changed.getvalue(), XLSX_CONTENT_TYPE)},
        )
        readiness = self.client.get("/api/readiness")

        self.assertEqual(activated.status_code, 201)
        self.assertFalse(readiness.json()["isReady"])

    def test_curriculum_mismatch_does_not_replace_active_version(self) -> None:
        self.upload("/api/imports/activate", "valid-import.xlsx")
        source = BytesIO((FIXTURES / "valid-import.xlsx").read_bytes())
        workbook = load_workbook(source)
        workbook["Нагрузка"]["D2"] = "UNKNOWN"
        changed = BytesIO()
        workbook.save(changed)
        workbook.close()

        rejected = self.client.post(
            "/api/imports/activate",
            files={"file": ("invalid-plan.xlsx", changed.getvalue(), XLSX_CONTENT_TYPE)},
        )
        status = self.client.get("/api/status")

        self.assertEqual(rejected.status_code, 422)
        self.assertEqual(
            rejected.json()["issues"][0]["code"],
            "discipline_not_in_curriculum",
        )
        self.assertEqual(status.json()["activeVersionId"], 1)
        self.assertEqual(len(status.json()["versions"]), 1)

    def test_preview_compares_students_with_active_version(self) -> None:
        self.upload("/api/imports/activate", "valid-import.xlsx")
        source = BytesIO((FIXTURES / "valid-import.xlsx").read_bytes())
        workbook = load_workbook(source)
        students = workbook["Студенты"]
        students["B2"] = "Петров Пётр Сергеевич"
        students.append(
            [
                "S-002",
                "Сидорова Анна Олеговна",
                "ИС-101",
                "active",
                "2026-09-01",
                None,
                "1",
                None,
            ]
        )
        changed = BytesIO()
        workbook.save(changed)
        workbook.close()

        preview = self.client.post(
            "/api/imports/preview",
            files={"file": ("students.xlsx", changed.getvalue(), XLSX_CONTENT_TYPE)},
        )

        self.assertEqual(preview.status_code, 200)
        self.assertEqual(
            preview.json()["studentChanges"],
            {"created": 1, "updated": 1, "deactivated": 0},
        )

    def test_saved_version_can_be_reactivated_and_unknown_version_is_404(self) -> None:
        self.upload("/api/imports/activate", "valid-import.xlsx")

        activated = self.client.post("/api/versions/1/activate")
        missing = self.client.post("/api/versions/404/activate")

        self.assertEqual(activated.status_code, 200)
        self.assertEqual(activated.json()["activeVersionId"], 1)
        self.assertEqual(missing.status_code, 404)
        self.assertEqual(missing.json()["error"]["code"], "version_not_found")

    def test_cross_site_mutation_is_rejected(self) -> None:
        response = self.upload(
            "/api/imports/activate",
            "valid-import.xlsx",
            headers={"Origin": "https://malicious.example"},
        )

        self.assertEqual(response.status_code, 403)
        self.assertFalse(self.database_path.exists())

    def test_oversized_upload_is_rejected_before_excel_parsing(self) -> None:
        response = self.client.post(
            "/api/imports/preview",
            files={
                "file": (
                    "large.xlsx",
                    b"0" * (10 * 1024 * 1024 + 1),
                    XLSX_CONTENT_TYPE,
                )
            },
        )

        self.assertEqual(response.status_code, 413)
        self.assertEqual(response.json()["error"]["code"], "file_too_large")


if __name__ == "__main__":
    unittest.main()
