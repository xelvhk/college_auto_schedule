from __future__ import annotations

import unittest
from datetime import date
from pathlib import Path
from tempfile import TemporaryDirectory

from openpyxl import Workbook, load_workbook
from pydantic import ValidationError

from rasp.domain.models import (
    Building,
    Equipment,
    Group,
    Room,
    RoomType,
    Student,
    Teacher,
    WorkloadItem,
)
from rasp.imports.excel import (
    RUSSIAN_HEADERS,
    ImportValidationError,
    build_import_batch,
    read_import_workbook,
)


FIXTURES = Path(__file__).parent / "fixtures"


class DomainModelTests(unittest.TestCase):
    def test_teacher_rejects_impossible_daily_limit(self) -> None:
        with self.assertRaises(ValidationError):
            Teacher(
                teacher_code="T-001",
                full_name="Иванова Ирина Игоревна",
                yearly_assigned_hours=720,
                max_hours_per_day=0,
            )

    def test_group_normalizes_code_and_requires_positive_headcount(self) -> None:
        group = Group(group_code="  ис-101  ", course=1, headcount=25)
        self.assertEqual(group.group_code, "ИС-101")

        with self.assertRaises(ValidationError):
            Group(group_code="ИС-102", course=1, headcount=0)

    def test_student_normalizes_memberships_and_orders_dates(self) -> None:
        student = Student(
            student_code=" s-001 ",
            full_name="Петров Пётр Петрович",
            group_code=" ис-101 ",
            status="active",
            enrollment_date="2026-09-01",
            subgroup_codes=(" 2 ", "1"),
            elective_codes=(" web ",),
        )

        self.assertEqual(student.student_code, "S-001")
        self.assertEqual(student.group_code, "ИС-101")
        self.assertEqual(student.subgroup_codes, (1, 2))
        self.assertEqual(student.elective_codes, ("WEB",))

        with self.assertRaises(ValidationError):
            Student(**(student.model_dump() | {"end_date": "2026-08-31"}))

    def test_room_normalizes_equipment_and_requires_positive_capacity(self) -> None:
        room = Room(
            room_code=" main-201 ",
            room_name="Лаборатория 201",
            building_code=" main ",
            room_type_code=" computer_lab ",
            capacity=25,
            equipment_codes=" projector ; computers ",
        )

        self.assertEqual(room.room_code, "MAIN-201")
        self.assertEqual(room.building_code, "MAIN")
        self.assertEqual(room.room_type_code, "COMPUTER_LAB")
        self.assertEqual(room.equipment_codes, ("COMPUTERS", "PROJECTOR"))

        with self.assertRaises(ValidationError):
            Room(**(room.model_dump() | {"capacity": 0}))


class ImportBatchTests(unittest.TestCase):
    def setUp(self) -> None:
        self.teacher_rows = [
            {
                "teacher_code": "T-001",
                "full_name": "Иванова Ирина Игоревна",
                "yearly_assigned_hours": 720,
                "max_hours_per_day": 8,
                "active": True,
            }
        ]
        self.group_rows = [
            {
                "group_code": "ИС-101",
                "course": 1,
                "headcount": 25,
                "subgroup_count": 1,
            }
        ]
        self.workload_rows = [
            {
                "workload_row_code": "W-001",
                "academic_year": "2026/2027",
                "semester": 1,
                "discipline_code": "MDK.01.01",
                "discipline_name": "Основы программирования",
                "group_code": "ИС-101",
                "teacher_code": "T-001",
                "lesson_type": "practice",
                "total_academic_hours": 72,
                "event_duration_hours": 2,
            }
        ]

    def test_builds_atomic_batch_when_rows_are_valid(self) -> None:
        batch = build_import_batch(
            teacher_rows=self.teacher_rows,
            group_rows=self.group_rows,
            workload_rows=self.workload_rows,
        )

        self.assertEqual(len(batch.teachers), 1)
        self.assertEqual(len(batch.groups), 1)
        self.assertEqual(len(batch.workloads), 1)
        self.assertEqual(batch.workloads[0].teacher_code, "T-001")

    def test_rejects_unknown_teacher_without_returning_partial_batch(self) -> None:
        self.workload_rows[0]["teacher_code"] = "T-404"

        with self.assertRaises(ImportValidationError) as raised:
            build_import_batch(
                teacher_rows=self.teacher_rows,
                group_rows=self.group_rows,
                workload_rows=self.workload_rows,
            )

        self.assertTrue(
            any(issue.code == "unknown_teacher" for issue in raised.exception.issues)
        )

    def test_rejects_duplicate_stable_codes(self) -> None:
        duplicate = dict(self.teacher_rows[0])
        duplicate["full_name"] = "Другой преподаватель"

        with self.assertRaises(ImportValidationError) as raised:
            build_import_batch(
                teacher_rows=[*self.teacher_rows, duplicate],
                group_rows=self.group_rows,
                workload_rows=self.workload_rows,
            )

        self.assertTrue(
            any(issue.code == "duplicate_code" for issue in raised.exception.issues)
        )

    def test_workload_requires_whole_number_of_events(self) -> None:
        with self.assertRaises(ValidationError):
            WorkloadItem(**{**self.workload_rows[0], "total_academic_hours": 71})

    def test_rejects_student_for_unknown_group_or_subgroup(self) -> None:
        student_rows = [
            {
                "student_code": "S-001",
                "full_name": "Петров Пётр Петрович",
                "group_code": "ИС-404",
                "status": "active",
                "subgroup_codes": "2",
            }
        ]

        with self.assertRaises(ImportValidationError) as raised:
            build_import_batch(
                teacher_rows=self.teacher_rows,
                group_rows=self.group_rows,
                workload_rows=self.workload_rows,
                student_rows=student_rows,
            )

        self.assertIn("unknown_student_group", {i.code for i in raised.exception.issues})

        student_rows[0]["group_code"] = "ИС-101"
        with self.assertRaises(ImportValidationError) as raised:
            build_import_batch(
                teacher_rows=self.teacher_rows,
                group_rows=self.group_rows,
                workload_rows=self.workload_rows,
                student_rows=student_rows,
            )

        self.assertIn("unknown_subgroup", {i.code for i in raised.exception.issues})

    def test_validates_room_references(self) -> None:
        common = {
            "teacher_rows": self.teacher_rows,
            "group_rows": self.group_rows,
            "workload_rows": self.workload_rows,
            "building_rows": [
                {"building_code": "MAIN", "building_name": "Главный корпус"}
            ],
            "room_type_rows": [
                {
                    "room_type_code": "COMPUTER_LAB",
                    "room_type_name": "Компьютерный класс",
                }
            ],
            "equipment_rows": [
                {"equipment_code": "COMPUTERS", "equipment_name": "Компьютеры"}
            ],
        }
        room_rows = [
            {
                "room_code": "MAIN-201",
                "room_name": "Лаборатория 201",
                "building_code": "MAIN",
                "room_type_code": "COMPUTER_LAB",
                "capacity": 25,
                "equipment_codes": "COMPUTERS",
            }
        ]

        batch = build_import_batch(**common, room_rows=room_rows)
        self.assertEqual(batch.rooms[0].equipment_codes, ("COMPUTERS",))

        room_rows[0]["building_code"] = "UNKNOWN"
        with self.assertRaises(ImportValidationError) as raised:
            build_import_batch(**common, room_rows=room_rows)

        self.assertIn("unknown_room_building", {i.code for i in raised.exception.issues})

    def test_workload_must_reference_imported_academic_year(self) -> None:
        with self.assertRaises(ImportValidationError) as raised:
            build_import_batch(
                teacher_rows=self.teacher_rows,
                group_rows=self.group_rows,
                workload_rows=self.workload_rows,
                academic_year_rows=[
                    {
                        "academic_year": "2025/2026",
                        "starts_on": "2025-09-01",
                        "ends_on": "2026-06-30",
                    }
                ],
                calendar_period_rows=[],
                bell_slot_rows=[],
            )

        self.assertIn("unknown_workload_academic_year", {i.code for i in raised.exception.issues})

    def test_unavailability_must_reference_resource_of_matching_type(self) -> None:
        calendar = {
            "academic_year_rows": [
                {
                    "academic_year": "2026/2027",
                    "starts_on": "2026-09-01",
                    "ends_on": "2027-06-30",
                }
            ],
            "calendar_period_rows": [],
            "bell_slot_rows": [],
            "calendar_exception_rows": [],
            "resource_unavailability_rows": [
                {
                    "unavailability_code": "U-001",
                    "academic_year": "2026/2027",
                    "resource_type": "room",
                    "resource_code": "T-001",
                    "starts_on": "2026-10-01",
                    "ends_on": "2026-10-01",
                }
            ],
        }

        with self.assertRaises(ImportValidationError) as raised:
            build_import_batch(
                teacher_rows=self.teacher_rows,
                group_rows=self.group_rows,
                workload_rows=self.workload_rows,
                **calendar,
            )

        self.assertIn(
            "unknown_unavailability_resource",
            {issue.code for issue in raised.exception.issues},
        )


class WorkbookImportTests(unittest.TestCase):
    def test_canonical_workbook_uses_russian_headers(self) -> None:
        workbook = load_workbook(FIXTURES / "valid-import.xlsx", read_only=True)
        try:
            self.assertEqual(workbook["Преподаватели"]["A1"].value, "Код преподавателя")
            self.assertEqual(workbook["Студенты"]["B1"].value, "ФИО")
            self.assertEqual(workbook["Аудитории"]["E1"].value, "Вместимость")
            self.assertEqual(
                workbook["Нагрузка"]["Q1"].value,
                "Требуемое оборудование",
            )
            self.assertEqual(workbook["Учебные годы"]["A1"].value, "Учебный год")
            self.assertEqual(workbook["Периоды"]["D1"].value, "Тип периода")
            self.assertEqual(workbook["Сетка звонков"]["E1"].value, "Начало")
            self.assertEqual(
                workbook["Исключения календаря"]["C1"].value,
                "Тип исключения",
            )
            self.assertEqual(
                workbook["Недоступность"]["D1"].value,
                "Код ресурса",
            )
        finally:
            workbook.close()

    def test_legacy_english_headers_remain_supported(self) -> None:
        workbook = load_workbook(FIXTURES / "valid-import.xlsx")
        for sheet_name, aliases in RUSSIAN_HEADERS.items():
            legacy_headers = {russian: canonical for canonical, russian in aliases.items()}
            for cell in workbook[sheet_name][1]:
                if cell.value in legacy_headers:
                    cell.value = legacy_headers[cell.value]

        with TemporaryDirectory() as directory:
            target = Path(directory) / "legacy-headers.xlsx"
            workbook.save(target)
            workbook.close()
            batch = read_import_workbook(target)

        self.assertEqual(batch.teachers[0].teacher_code, "T-001")
        self.assertEqual(batch.rooms[0].room_code, "MAIN-201")
        self.assertEqual(batch.academic_years[0].academic_year, "2026/2027")
        self.assertEqual(batch.calendar_periods[0].period_code, "SEM-1")
        self.assertEqual(batch.bell_slots[0].slot_code, "S1-01")
        self.assertEqual(batch.calendar_exceptions[0].exception_code, "EX-001")
        self.assertEqual(
            batch.resource_unavailability[0].unavailability_code,
            "U-001",
        )

    def test_legacy_calendar_without_constraint_sheets_remains_supported(self) -> None:
        workbook = load_workbook(FIXTURES / "valid-import.xlsx")
        del workbook["Исключения календаря"]
        del workbook["Недоступность"]

        with TemporaryDirectory() as directory:
            target = Path(directory) / "legacy-calendar.xlsx"
            workbook.save(target)
            workbook.close()
            batch = read_import_workbook(target)

        self.assertEqual(len(batch.academic_years), 1)
        self.assertEqual(batch.calendar_exceptions, ())
        self.assertEqual(batch.resource_unavailability, ())

    def test_rejects_partial_calendar_sheet_set(self) -> None:
        workbook = load_workbook(FIXTURES / "valid-import.xlsx")
        del workbook["Сетка звонков"]

        with TemporaryDirectory() as directory:
            target = Path(directory) / "partial-calendar.xlsx"
            workbook.save(target)
            workbook.close()
            with self.assertRaises(ImportValidationError) as raised:
                read_import_workbook(target)

        self.assertEqual(raised.exception.issues[0].code, "missing_sheet")
        self.assertIn("Сетка звонков", raised.exception.issues[0].message)

    def test_rejects_partial_calendar_constraint_sheet_set(self) -> None:
        workbook = load_workbook(FIXTURES / "valid-import.xlsx")
        del workbook["Недоступность"]

        with TemporaryDirectory() as directory:
            target = Path(directory) / "partial-constraints.xlsx"
            workbook.save(target)
            workbook.close()
            with self.assertRaises(ImportValidationError) as raised:
                read_import_workbook(target)

        self.assertEqual(raised.exception.issues[0].code, "missing_sheet")
        self.assertIn("Недоступность", raised.exception.issues[0].message)

    def test_rejects_forbidden_personal_columns(self) -> None:
        workbook = load_workbook(FIXTURES / "valid-import.xlsx")
        students = workbook["Студенты"]
        students["I1"] = "Паспортные данные"
        students["I2"] = "x"

        with TemporaryDirectory() as directory:
            target = Path(directory) / "sensitive.xlsx"
            workbook.save(target)
            workbook.close()
            with self.assertRaises(ImportValidationError) as raised:
                read_import_workbook(target)

        self.assertIn("forbidden_personal_data", {i.code for i in raised.exception.issues})

    def test_rejects_partial_reference_sheet_set(self) -> None:
        workbook = load_workbook(FIXTURES / "valid-import.xlsx")
        del workbook["Дисциплины"]

        with TemporaryDirectory() as directory:
            target = Path(directory) / "partial.xlsx"
            workbook.save(target)
            workbook.close()
            with self.assertRaises(ImportValidationError) as raised:
                read_import_workbook(target)

        self.assertEqual(raised.exception.issues[0].code, "missing_sheet")
        self.assertIn("Дисциплины", raised.exception.issues[0].message)

    def test_reads_reference_sheets_into_the_same_batch(self) -> None:
        workbook = Workbook()
        workbook.remove(workbook.active)
        sheets = {
            "Преподаватели": (
                ["teacher_code", "full_name", "yearly_assigned_hours"],
                [["T-001", "Иванова Ирина Игоревна", 72]],
            ),
            "Группы": (
                [
                    "group_code",
                    "specialty_code",
                    "curriculum_code",
                    "course",
                    "headcount",
                ],
                [["ИС-101", "09.02.07", "UP-09.02.07-2026", 1, 25]],
            ),
            "Нагрузка": (
                [
                    "workload_row_code",
                    "academic_year",
                    "semester",
                    "discipline_code",
                    "discipline_name",
                    "group_code",
                    "teacher_code",
                    "lesson_type",
                    "total_academic_hours",
                    "event_duration_hours",
                ],
                [[
                    "W-001", "2026/2027", 1, "MDK.01.01",
                    "Основы программирования", "ИС-101", "T-001",
                    "practice", 72, 2,
                ]],
            ),
            "Специальности": (
                [
                    "specialty_code",
                    "specialty_name",
                    "program_base",
                    "education_form",
                ],
                [["09.02.07", "Информационные системы", "9", "full_time"]],
            ),
            "Учебные планы": (
                [
                    "curriculum_code",
                    "specialty_code",
                    "admission_year",
                    "version",
                    "valid_from",
                    "status",
                ],
                [[
                    "UP-09.02.07-2026", "09.02.07", 2026, "1.0",
                    date(2026, 9, 1), "active",
                ]],
            ),
            "Дисциплины": (
                [
                    "curriculum_code",
                    "discipline_code",
                    "discipline_name",
                    "semester",
                    "lesson_type",
                    "planned_hours",
                ],
                [[
                    "UP-09.02.07-2026", "MDK.01.01",
                    "Основы программирования", 1, "practice", 72,
                ]],
            ),
        }
        for title, (headers, rows) in sheets.items():
            worksheet = workbook.create_sheet(title)
            worksheet.append(headers)
            for row in rows:
                worksheet.append(row)

        with TemporaryDirectory() as directory:
            target = Path(directory) / "full.xlsx"
            workbook.save(target)
            workbook.close()
            batch = read_import_workbook(target)

        self.assertEqual(batch.specialties[0].specialty_code, "09.02.07")
        self.assertEqual(batch.curricula[0].curriculum_code, "UP-09.02.07-2026")
        self.assertEqual(batch.disciplines[0].planned_hours, 72)

    def test_reads_valid_canonical_workbook(self) -> None:
        batch = read_import_workbook(FIXTURES / "valid-import.xlsx")

        self.assertEqual(batch.teachers[0].teacher_code, "T-001")
        self.assertEqual(batch.groups[0].group_code, "ИС-101")
        self.assertEqual(batch.workloads[0].workload_row_code, "W-001")
        self.assertEqual(batch.specialties[0].specialty_code, "09.02.07")
        self.assertEqual(batch.students[0].student_code, "S-001")
        self.assertEqual(batch.buildings[0].building_code, "MAIN")
        self.assertEqual(batch.room_types[0].room_type_code, "COMPUTER_LAB")
        self.assertEqual(batch.equipment[0].equipment_code, "COMPUTERS")
        self.assertEqual(batch.rooms[0].room_code, "MAIN-201")

    def test_rejects_non_xlsx_file_before_parsing(self) -> None:
        with self.assertRaises(ImportValidationError) as raised:
            read_import_workbook(__file__)

        self.assertEqual(raised.exception.issues[0].code, "invalid_extension")

    def test_rejects_formula_cells(self) -> None:
        with self.assertRaises(ImportValidationError) as raised:
            read_import_workbook(FIXTURES / "formula-import.xlsx")

        self.assertTrue(
            any(issue.code == "formula_forbidden" for issue in raised.exception.issues)
        )


if __name__ == "__main__":
    unittest.main()
