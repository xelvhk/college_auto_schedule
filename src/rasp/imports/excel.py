from __future__ import annotations

from collections.abc import Iterable, Mapping
from dataclasses import dataclass
from pathlib import Path, PurePosixPath
from typing import Any, TypeVar
from zipfile import BadZipFile, ZipFile, is_zipfile

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import BaseModel, ValidationError

from rasp.domain.models import (
    Building,
    Equipment,
    Group,
    ImportBatch,
    ReferenceDataBatch,
    Room,
    RoomType,
    Student,
    Teacher,
    WorkloadItem,
)
from rasp.imports.reference_data import (
    ReferenceDataValidationError,
    build_reference_data_batch,
)
from rasp.imports.calendar import CalendarValidationError, build_calendar_batch


@dataclass(frozen=True, slots=True)
class ImportIssue:
    section: str
    row: int
    column: str | None
    code: str
    message: str


class ImportValidationError(ValueError):
    def __init__(self, issues: Iterable[ImportIssue]) -> None:
        self.issues = tuple(issues)
        super().__init__(f"Import contains {len(self.issues)} validation issue(s)")


ModelT = TypeVar("ModelT", bound=BaseModel)

MAX_FILE_SIZE = 10 * 1024 * 1024
MAX_ARCHIVE_ENTRIES = 1_000
MAX_UNCOMPRESSED_SIZE = 50 * 1024 * 1024
MAX_COMPRESSION_RATIO = 100
MAX_DATA_ROWS_PER_SHEET = 100_000

CORE_SHEETS = {
    "Преподаватели": "teachers",
    "Группы": "groups",
    "Нагрузка": "workloads",
}
REFERENCE_SHEETS = {
    "Специальности": "specialties",
    "Учебные планы": "curricula",
    "Дисциплины": "disciplines",
}
STUDENT_SHEETS = {"Студенты": "students"}
ROOM_SHEETS = {
    "Корпуса": "buildings",
    "Типы помещений": "room_types",
    "Оборудование": "equipment",
    "Аудитории": "rooms",
}
CALENDAR_SHEETS = {
    "Учебные годы": "academic_years",
    "Периоды": "calendar_periods",
    "Сетка звонков": "bell_slots",
}
CALENDAR_CONSTRAINT_SHEETS = {
    "Исключения календаря": "calendar_exceptions",
    "Недоступность": "resource_unavailability",
}
RUSSIAN_HEADERS = {
    "Преподаватели": {
        "teacher_code": "Код преподавателя",
        "full_name": "ФИО",
        "department": "Подразделение",
        "employment_type": "Тип занятости",
        "yearly_assigned_hours": "Назначено часов в год",
        "yearly_limit_hours": "Лимит часов в год",
        "max_hours_per_day": "Макс. часов в день",
        "max_days_per_week": "Макс. дней в неделю",
        "home_building_code": "Основной корпус",
        "active": "Активен",
    },
    "Группы": {
        "group_code": "Код группы",
        "specialty_code": "Код специальности",
        "curriculum_code": "Код учебного плана",
        "course": "Курс",
        "education_form": "Форма обучения",
        "headcount": "Численность",
        "program_base": "База обучения",
        "study_week_type": "Тип учебной недели",
        "primary_building_code": "Основной корпус",
        "subgroup_count": "Количество подгрупп",
    },
    "Специальности": {
        "specialty_code": "Код специальности",
        "specialty_name": "Название специальности",
        "qualification": "Квалификация",
        "program_base": "База обучения",
        "education_form": "Форма обучения",
        "active": "Активна",
    },
    "Учебные планы": {
        "curriculum_code": "Код учебного плана",
        "specialty_code": "Код специальности",
        "admission_year": "Год набора",
        "version": "Версия",
        "valid_from": "Действует с",
        "valid_to": "Действует по",
        "status": "Статус",
    },
    "Дисциплины": {
        "curriculum_code": "Код учебного плана",
        "discipline_code": "Код дисциплины",
        "discipline_name": "Название дисциплины",
        "section_code": "Код раздела",
        "semester": "Семестр",
        "lesson_type": "Вид занятия",
        "planned_hours": "Плановые часы",
        "control_form": "Форма контроля",
    },
    "Студенты": {
        "student_code": "Код студента",
        "full_name": "ФИО",
        "group_code": "Код группы",
        "status": "Статус",
        "enrollment_date": "Дата зачисления",
        "end_date": "Дата окончания",
        "subgroup_codes": "Номера подгрупп",
        "elective_codes": "Коды элективов",
    },
    "Корпуса": {
        "building_code": "Код корпуса",
        "building_name": "Название корпуса",
        "active": "Активен",
    },
    "Типы помещений": {
        "room_type_code": "Код типа помещения",
        "room_type_name": "Название типа помещения",
        "active": "Активен",
    },
    "Оборудование": {
        "equipment_code": "Код оборудования",
        "equipment_name": "Название оборудования",
        "active": "Активно",
    },
    "Аудитории": {
        "room_code": "Код аудитории",
        "room_name": "Название аудитории",
        "building_code": "Код корпуса",
        "room_type_code": "Код типа помещения",
        "capacity": "Вместимость",
        "equipment_codes": "Коды оборудования",
        "active": "Активна",
    },
    "Нагрузка": {
        "workload_row_code": "Код строки нагрузки",
        "academic_year": "Учебный год",
        "semester": "Семестр",
        "discipline_code": "Код дисциплины",
        "discipline_name": "Название дисциплины",
        "group_code": "Код группы",
        "subgroup": "Подгруппа",
        "stream": "Поток",
        "teacher_code": "Код преподавателя",
        "lesson_type": "Вид занятия",
        "total_academic_hours": "Всего академических часов",
        "event_duration_hours": "Продолжительность занятия",
        "recurrence": "Периодичность",
        "lesson_bundle_code": "Код связки занятий",
        "room_type": "Тип помещения",
        "room_capacity": "Требуемая вместимость",
        "required_equipment_codes": "Требуемое оборудование",
    },
    "Учебные годы": {
        "academic_year": "Учебный год",
        "starts_on": "Дата начала",
        "ends_on": "Дата окончания",
        "active": "Активен",
    },
    "Периоды": {
        "period_code": "Код периода",
        "academic_year": "Учебный год",
        "period_name": "Название периода",
        "period_type": "Тип периода",
        "starts_on": "Дата начала",
        "ends_on": "Дата окончания",
        "semester": "Семестр",
    },
    "Сетка звонков": {
        "slot_code": "Код интервала",
        "academic_year": "Учебный год",
        "shift_code": "Код смены",
        "lesson_number": "Номер занятия",
        "starts_at": "Начало",
        "ends_at": "Окончание",
    },
    "Исключения календаря": {
        "exception_code": "Код исключения",
        "academic_year": "Учебный год",
        "exception_type": "Тип исключения",
        "exception_date": "Дата исключения",
        "transferred_to": "Перенос на дату",
        "shortened_ends_at": "Сокращённый день до",
        "note": "Примечание",
    },
    "Недоступность": {
        "unavailability_code": "Код недоступности",
        "academic_year": "Учебный год",
        "resource_type": "Тип ресурса",
        "resource_code": "Код ресурса",
        "starts_on": "Дата начала",
        "ends_on": "Дата окончания",
        "starts_at": "Время начала",
        "ends_at": "Время окончания",
        "reason": "Причина",
    },
}
FORBIDDEN_STUDENT_HEADERS = {
    "address",
    "attendance",
    "birth_date",
    "date_of_birth",
    "grades",
    "health_data",
    "medical_data",
    "passport",
    "passport_data",
    "payment_data",
    "phone",
    "адрес",
    "данные об оплате",
    "дата рождения",
    "медицинские данные",
    "номер телефона",
    "оценки",
    "паспорт",
    "паспортные данные",
    "посещаемость",
    "сведения о здоровье",
    "телефон",
}
REQUIRED_HEADERS = {
    "Преподаватели": {"teacher_code", "full_name", "yearly_assigned_hours"},
    "Группы": {"group_code", "course", "headcount"},
    "Нагрузка": {
        "workload_row_code",
        "academic_year",
        "semester",
        "discipline_code",
        "discipline_name",
        "group_code",
        "teacher_code",
        "lesson_type",
        "total_academic_hours",
        "event_duration_hours",
    },
    "Специальности": {
        "specialty_code",
        "specialty_name",
        "program_base",
        "education_form",
    },
    "Учебные планы": {
        "curriculum_code",
        "specialty_code",
        "admission_year",
        "version",
        "valid_from",
        "status",
    },
    "Дисциплины": {
        "curriculum_code",
        "discipline_code",
        "discipline_name",
        "semester",
        "lesson_type",
        "planned_hours",
    },
    "Студенты": {"student_code", "full_name", "group_code", "status"},
    "Корпуса": {"building_code", "building_name"},
    "Типы помещений": {"room_type_code", "room_type_name"},
    "Оборудование": {"equipment_code", "equipment_name"},
    "Аудитории": {
        "room_code",
        "room_name",
        "building_code",
        "room_type_code",
        "capacity",
    },
    "Учебные годы": {"academic_year", "starts_on", "ends_on"},
    "Периоды": {
        "period_code", "academic_year", "period_name", "period_type",
        "starts_on", "ends_on",
    },
    "Сетка звонков": {
        "slot_code", "academic_year", "shift_code", "lesson_number",
        "starts_at", "ends_at",
    },
    "Исключения календаря": {
        "exception_code", "academic_year", "exception_type", "exception_date",
    },
    "Недоступность": {
        "unavailability_code", "academic_year", "resource_type", "resource_code",
        "starts_on", "ends_on",
    },
}


def _parse_rows(
    section: str,
    rows: Iterable[Mapping[str, Any]],
    model: type[ModelT],
) -> tuple[list[ModelT], list[ImportIssue]]:
    parsed: list[ModelT] = []
    issues: list[ImportIssue] = []

    for row_number, row in enumerate(rows, start=2):
        try:
            parsed.append(model.model_validate(dict(row)))
        except ValidationError as error:
            for detail in error.errors(include_url=False, include_input=False):
                location = detail.get("loc", ())
                issues.append(
                    ImportIssue(
                        section=section,
                        row=row_number,
                        column=str(location[0]) if location else None,
                        code="invalid_value",
                        message=str(detail["msg"]),
                    )
                )

    return parsed, issues


def _duplicate_issues(
    section: str,
    records: Iterable[BaseModel],
    field: str,
) -> list[ImportIssue]:
    issues: list[ImportIssue] = []
    seen: set[str] = set()
    for row_number, record in enumerate(records, start=2):
        code = str(getattr(record, field))
        if code in seen:
            issues.append(
                ImportIssue(
                    section=section,
                    row=row_number,
                    column=field,
                    code="duplicate_code",
                    message=f"Duplicate stable code: {code}",
                )
            )
        seen.add(code)
    return issues


def build_import_batch(
    *,
    teacher_rows: Iterable[Mapping[str, Any]],
    group_rows: Iterable[Mapping[str, Any]],
    workload_rows: Iterable[Mapping[str, Any]],
    specialty_rows: Iterable[Mapping[str, Any]] | None = None,
    curriculum_rows: Iterable[Mapping[str, Any]] | None = None,
    discipline_rows: Iterable[Mapping[str, Any]] | None = None,
    student_rows: Iterable[Mapping[str, Any]] | None = None,
    building_rows: Iterable[Mapping[str, Any]] | None = None,
    room_type_rows: Iterable[Mapping[str, Any]] | None = None,
    equipment_rows: Iterable[Mapping[str, Any]] | None = None,
    room_rows: Iterable[Mapping[str, Any]] | None = None,
    academic_year_rows: Iterable[Mapping[str, Any]] | None = None,
    calendar_period_rows: Iterable[Mapping[str, Any]] | None = None,
    bell_slot_rows: Iterable[Mapping[str, Any]] | None = None,
    calendar_exception_rows: Iterable[Mapping[str, Any]] | None = None,
    resource_unavailability_rows: Iterable[Mapping[str, Any]] | None = None,
) -> ImportBatch:
    """Validate all rows and return a complete batch or no batch at all."""

    teachers, issues = _parse_rows("teachers", teacher_rows, Teacher)
    groups, group_issues = _parse_rows("groups", group_rows, Group)
    workloads, workload_issues = _parse_rows(
        "workloads", workload_rows, WorkloadItem
    )
    issues.extend(group_issues)
    issues.extend(workload_issues)
    students: list[Student] = []
    if student_rows is not None:
        students, student_issues = _parse_rows("students", student_rows, Student)
        issues.extend(student_issues)
    buildings: list[Building] = []
    room_types: list[RoomType] = []
    equipment: list[Equipment] = []
    rooms: list[Room] = []
    room_inputs = (building_rows, room_type_rows, equipment_rows, room_rows)
    if any(rows is not None for rows in room_inputs):
        if not all(rows is not None for rows in room_inputs):
            issues.append(
                ImportIssue(
                    "file", 0, None, "incomplete_room_data",
                    "All room reference sections must be provided together",
                )
            )
        else:
            buildings, parsed = _parse_rows("buildings", building_rows or (), Building)
            issues.extend(parsed)
            room_types, parsed = _parse_rows("room_types", room_type_rows or (), RoomType)
            issues.extend(parsed)
            equipment, parsed = _parse_rows("equipment", equipment_rows or (), Equipment)
            issues.extend(parsed)
            rooms, parsed = _parse_rows("rooms", room_rows or (), Room)
            issues.extend(parsed)

    issues.extend(_duplicate_issues("teachers", teachers, "teacher_code"))
    issues.extend(_duplicate_issues("groups", groups, "group_code"))
    issues.extend(
        _duplicate_issues("workloads", workloads, "workload_row_code")
    )
    issues.extend(_duplicate_issues("students", students, "student_code"))
    issues.extend(_duplicate_issues("buildings", buildings, "building_code"))
    issues.extend(_duplicate_issues("room_types", room_types, "room_type_code"))
    issues.extend(_duplicate_issues("equipment", equipment, "equipment_code"))
    issues.extend(_duplicate_issues("rooms", rooms, "room_code"))

    teacher_codes = {teacher.teacher_code for teacher in teachers}
    group_codes = {group.group_code for group in groups}
    for row_number, workload in enumerate(workloads, start=2):
        if workload.teacher_code not in teacher_codes:
            issues.append(
                ImportIssue(
                    section="workloads",
                    row=row_number,
                    column="teacher_code",
                    code="unknown_teacher",
                    message=f"Unknown teacher code: {workload.teacher_code}",
                )
            )
        if workload.group_code not in group_codes:
            issues.append(
                ImportIssue(
                    section="workloads",
                    row=row_number,
                    column="group_code",
                    code="unknown_group",
                    message=f"Unknown group code: {workload.group_code}",
                )
            )

    groups_by_code = {group.group_code: group for group in groups}
    for row_number, student in enumerate(students, start=2):
        group = groups_by_code.get(student.group_code)
        if group is None:
            issues.append(
                ImportIssue(
                    "students",
                    row_number,
                    "group_code",
                    "unknown_student_group",
                    f"Unknown student group code: {student.group_code}",
                )
            )
            continue
        for subgroup_code in student.subgroup_codes:
            if subgroup_code < 1 or subgroup_code > group.subgroup_count:
                issues.append(
                    ImportIssue(
                        "students",
                        row_number,
                        "subgroup_codes",
                        "unknown_subgroup",
                        "Student subgroup is outside the group's configured range",
                    )
                )

    building_codes = {item.building_code for item in buildings}
    room_type_codes = {item.room_type_code for item in room_types}
    equipment_codes = {item.equipment_code for item in equipment}
    if buildings:
        for row_number, teacher in enumerate(teachers, start=2):
            if (
                teacher.home_building_code
                and teacher.home_building_code not in building_codes
            ):
                issues.append(
                    ImportIssue(
                        "teachers",
                        row_number,
                        "home_building_code",
                        "unknown_teacher_building",
                        "Teacher references an unknown home building",
                    )
                )
        for row_number, group in enumerate(groups, start=2):
            if (
                group.primary_building_code
                and group.primary_building_code not in building_codes
            ):
                issues.append(
                    ImportIssue(
                        "groups",
                        row_number,
                        "primary_building_code",
                        "unknown_group_building",
                        "Group references an unknown primary building",
                    )
                )
    for row_number, room in enumerate(rooms, start=2):
        if room.building_code not in building_codes:
            issues.append(
                ImportIssue(
                    "rooms",
                    row_number,
                    "building_code",
                    "unknown_room_building",
                    "Room references an unknown building",
                )
            )
        if room.room_type_code not in room_type_codes:
            issues.append(
                ImportIssue(
                    "rooms",
                    row_number,
                    "room_type_code",
                    "unknown_room_type",
                    "Room references an unknown room type",
                )
            )
        if unknown := set(room.equipment_codes) - equipment_codes:
            issues.append(
                ImportIssue(
                    "rooms",
                    row_number,
                    "equipment_codes",
                    "unknown_room_equipment",
                    f"Room references unknown equipment codes: {', '.join(sorted(unknown))}",
                )
            )
    if room_types:
        for row_number, workload in enumerate(workloads, start=2):
            if workload.room_type and workload.room_type not in room_type_codes:
                issues.append(
                    ImportIssue(
                        "workloads",
                        row_number,
                        "room_type",
                        "unknown_required_room_type",
                        "Workload references an unknown room type",
                    )
                )
            if unknown := set(workload.required_equipment_codes) - equipment_codes:
                unknown_codes = ", ".join(sorted(unknown))
                issues.append(
                    ImportIssue(
                        "workloads",
                        row_number,
                        "required_equipment_codes",
                        "unknown_required_equipment",
                        f"Workload references unknown equipment codes: {unknown_codes}",
                    )
                )
    references = ReferenceDataBatch(specialties=(), curricula=(), disciplines=())
    reference_inputs = (specialty_rows, curriculum_rows, discipline_rows)
    if any(rows is not None for rows in reference_inputs):
        if not all(rows is not None for rows in reference_inputs):
            issues.append(
                ImportIssue(
                    section="file",
                    row=0,
                    column=None,
                    code="incomplete_reference_data",
                    message="All curriculum reference sections must be provided together",
                )
            )
        else:
            try:
                references = build_reference_data_batch(
                    specialty_rows=specialty_rows or (),
                    curriculum_rows=curriculum_rows or (),
                    discipline_rows=discipline_rows or (),
                )
            except ReferenceDataValidationError as error:
                issues.extend(
                    ImportIssue(
                        section=issue.section,
                        row=issue.row,
                        column=issue.column,
                        code=issue.code,
                        message=issue.message,
                    )
                    for issue in error.issues
                )

    calendar = None
    base_calendar_inputs = (
        academic_year_rows,
        calendar_period_rows,
        bell_slot_rows,
    )
    constraint_inputs = (
        calendar_exception_rows,
        resource_unavailability_rows,
    )
    has_base_calendar = any(rows is not None for rows in base_calendar_inputs)
    has_constraints = any(rows is not None for rows in constraint_inputs)
    if has_base_calendar or has_constraints:
        if not all(rows is not None for rows in base_calendar_inputs):
            issues.append(
                ImportIssue(
                    "file", 0, None, "incomplete_calendar_data",
                    "All base calendar sections must be provided together",
                )
            )
        elif has_constraints and not all(
            rows is not None for rows in constraint_inputs
        ):
            issues.append(
                ImportIssue(
                    "file", 0, None, "incomplete_calendar_constraints",
                    "Both calendar constraint sections must be provided together",
                )
            )
        else:
            try:
                calendar = build_calendar_batch(
                    academic_year_rows=academic_year_rows or (),
                    period_rows=calendar_period_rows or (),
                    bell_slot_rows=bell_slot_rows or (),
                    exception_rows=calendar_exception_rows or (),
                    unavailability_rows=resource_unavailability_rows or (),
                )
            except CalendarValidationError as error:
                issues.extend(
                    ImportIssue(
                        issue.section, issue.row, issue.column, issue.code, issue.message
                    )
                    for issue in error.issues
                )
    if calendar is not None:
        academic_year_codes = {
            item.academic_year for item in calendar.academic_years
        }
        for row_number, workload in enumerate(workloads, start=2):
            if workload.academic_year not in academic_year_codes:
                issues.append(
                    ImportIssue(
                        "workloads",
                        row_number,
                        "academic_year",
                        "unknown_workload_academic_year",
                        "Workload references an unknown academic year",
                    )
                )
        resource_codes = {
            "teacher": {item.teacher_code for item in teachers},
            "group": {item.group_code for item in groups},
            "room": {item.room_code for item in rooms},
        }
        for row_number, unavailable in enumerate(calendar.unavailability, start=2):
            if unavailable.resource_code not in resource_codes[
                unavailable.resource_type.value
            ]:
                issues.append(
                    ImportIssue(
                        "resource_unavailability",
                        row_number,
                        "resource_code",
                        "unknown_unavailability_resource",
                        "Unavailability references an unknown resource of this type",
                    )
                )

    operational_batch = ImportBatch(
        teachers=tuple(teachers),
        groups=tuple(groups),
        workloads=tuple(workloads),
        students=tuple(students),
        buildings=tuple(buildings),
        room_types=tuple(room_types),
        equipment=tuple(equipment),
        rooms=tuple(rooms),
    )
    if issues:
        raise ImportValidationError(issues)

    return operational_batch.model_copy(
        update={
            "specialties": references.specialties,
            "curricula": references.curricula,
            "disciplines": references.disciplines,
            "academic_years": calendar.academic_years if calendar else (),
            "calendar_periods": calendar.periods if calendar else (),
            "bell_slots": calendar.bell_slots if calendar else (),
            "calendar_exceptions": calendar.exceptions if calendar else (),
            "resource_unavailability": calendar.unavailability if calendar else (),
        }
    )


def preflight_import_workbook(path: str | Path) -> None:
    path = Path(path)
    issues: list[ImportIssue] = []
    if path.suffix.lower() != ".xlsx":
        issues.append(
            ImportIssue("file", 0, None, "invalid_extension", "Expected .xlsx file")
        )
    if not path.is_file():
        issues.append(
            ImportIssue("file", 0, None, "file_not_found", "File does not exist")
        )
    elif path.stat().st_size > MAX_FILE_SIZE:
        issues.append(
            ImportIssue("file", 0, None, "file_too_large", "File exceeds 10 MiB")
        )
    if issues:
        raise ImportValidationError(issues)
    if not is_zipfile(path):
        raise ImportValidationError(
            [ImportIssue("file", 0, None, "invalid_xlsx", "Invalid XLSX container")]
        )

    try:
        with ZipFile(path) as archive:
            entries = archive.infolist()
            if len(entries) > MAX_ARCHIVE_ENTRIES:
                issues.append(
                    ImportIssue(
                        "file", 0, None, "archive_too_large", "Too many ZIP entries"
                    )
                )

            total_size = 0
            for entry in entries:
                archive_path = PurePosixPath(entry.filename)
                if archive_path.is_absolute() or ".." in archive_path.parts:
                    issues.append(
                        ImportIssue(
                            "file",
                            0,
                            None,
                            "unsafe_archive_path",
                            "Unsafe path inside XLSX container",
                        )
                    )
                if entry.flag_bits & 0x1:
                    issues.append(
                        ImportIssue(
                            "file",
                            0,
                            None,
                            "encrypted_content",
                            "Encrypted XLSX content is not supported",
                        )
                    )
                total_size += entry.file_size
                ratio = entry.file_size / max(entry.compress_size, 1)
                if ratio > MAX_COMPRESSION_RATIO:
                    issues.append(
                        ImportIssue(
                            "file",
                            0,
                            None,
                            "suspicious_compression",
                            "Suspicious ZIP compression ratio",
                        )
                    )
                normalized_name = entry.filename.lower()
                if (
                    "vbaproject.bin" in normalized_name
                    or normalized_name.startswith("xl/externallinks/")
                    or normalized_name.startswith("xl/embeddings/")
                ):
                    issues.append(
                        ImportIssue(
                            "file",
                            0,
                            None,
                            "active_content",
                            "Macros, external links, and embedded objects are forbidden",
                        )
                    )
            if total_size > MAX_UNCOMPRESSED_SIZE:
                issues.append(
                    ImportIssue(
                        "file",
                        0,
                        None,
                        "archive_too_large",
                        "Uncompressed XLSX content exceeds 50 MiB",
                    )
                )
    except BadZipFile as error:
        raise ImportValidationError(
            [ImportIssue("file", 0, None, "invalid_xlsx", "Invalid XLSX container")]
        ) from error

    if issues:
        raise ImportValidationError(issues)


def _read_sheet_rows(worksheet: Any) -> list[dict[str, Any]]:
    issues: list[ImportIssue] = []
    row_iterator = worksheet.iter_rows(values_only=False)
    header_cells = next(row_iterator, None)
    if header_cells is None:
        raise ImportValidationError(
            [
                ImportIssue(
                    worksheet.title, 1, None, "empty_sheet", "Sheet is empty"
                )
            ]
        )

    headers: list[str] = []
    for cell in header_cells:
        if cell.data_type == "f":
            issues.append(
                ImportIssue(
                    worksheet.title,
                    1,
                    cell.coordinate,
                    "formula_forbidden",
                    "Formulas are forbidden in import files",
                )
            )
        value = cell.value
        headers.append(str(value).strip() if value is not None else "")

    while headers and not headers[-1]:
        headers.pop()
    if not headers or any(not header for header in headers):
        issues.append(
            ImportIssue(
                worksheet.title,
                1,
                None,
                "invalid_header",
                "Header cells must be non-empty",
            )
        )
    aliases = {
        russian.casefold(): canonical
        for canonical, russian in RUSSIAN_HEADERS[worksheet.title].items()
    }
    headers = [aliases.get(header.casefold(), header) for header in headers]
    duplicate_headers = {header for header in headers if headers.count(header) > 1}
    if duplicate_headers:
        issues.append(
            ImportIssue(
                worksheet.title,
                1,
                None,
                "duplicate_header",
                f"Duplicate headers: {', '.join(sorted(duplicate_headers))}",
            )
        )
    missing = REQUIRED_HEADERS[worksheet.title] - set(headers)
    if missing:
        issues.append(
            ImportIssue(
                worksheet.title,
                1,
                None,
                "missing_header",
                f"Missing headers: {', '.join(sorted(missing))}",
            )
        )
    if worksheet.title == "Студенты":
        forbidden = FORBIDDEN_STUDENT_HEADERS & {header.lower() for header in headers}
        if forbidden:
            issues.append(
                ImportIssue(
                    worksheet.title,
                    1,
                    None,
                    "forbidden_personal_data",
                    f"Forbidden personal-data headers: {', '.join(sorted(forbidden))}",
                )
            )

    rows: list[dict[str, Any]] = []
    for row_number, cells in enumerate(row_iterator, start=2):
        if row_number > MAX_DATA_ROWS_PER_SHEET + 1:
            issues.append(
                ImportIssue(
                    worksheet.title,
                    row_number,
                    None,
                    "too_many_rows",
                    "Sheet exceeds 100000 data rows",
                )
            )
            break
        values = []
        for cell in cells[: len(headers)]:
            if cell.data_type == "f":
                issues.append(
                    ImportIssue(
                        worksheet.title,
                        row_number,
                        cell.coordinate,
                        "formula_forbidden",
                        "Formulas are forbidden in import files",
                    )
                )
            values.append(cell.value)
        values.extend([None] * (len(headers) - len(values)))
        if any(value is not None and value != "" for value in values):
            rows.append(dict(zip(headers, values, strict=True)))

    if issues:
        raise ImportValidationError(issues)
    return rows


def read_import_workbook(path: str | Path) -> ImportBatch:
    """Read the required sheets and return a fully validated, in-memory batch."""

    source = Path(path)
    preflight_import_workbook(source)
    try:
        workbook = load_workbook(
            source,
            read_only=True,
            data_only=False,
            keep_links=False,
        )
    except (InvalidFileException, OSError, ValueError) as error:
        raise ImportValidationError(
            [ImportIssue("file", 0, None, "invalid_xlsx", "Cannot read XLSX file")]
        ) from error

    try:
        available_sheets = set(workbook.sheetnames)
        missing_sheets = set(CORE_SHEETS) - available_sheets
        reference_sheets_present = set(REFERENCE_SHEETS) & available_sheets
        room_sheets_present = set(ROOM_SHEETS) & available_sheets
        calendar_sheets_present = set(CALENDAR_SHEETS) & available_sheets
        constraint_sheets_present = set(CALENDAR_CONSTRAINT_SHEETS) & available_sheets
        if reference_sheets_present and reference_sheets_present != set(
            REFERENCE_SHEETS
        ):
            missing_sheets |= set(REFERENCE_SHEETS) - available_sheets
        if room_sheets_present and room_sheets_present != set(ROOM_SHEETS):
            missing_sheets |= set(ROOM_SHEETS) - available_sheets
        if calendar_sheets_present and calendar_sheets_present != set(CALENDAR_SHEETS):
            missing_sheets |= set(CALENDAR_SHEETS) - available_sheets
        if constraint_sheets_present and constraint_sheets_present != set(
            CALENDAR_CONSTRAINT_SHEETS
        ):
            missing_sheets |= set(CALENDAR_CONSTRAINT_SHEETS) - available_sheets
        if constraint_sheets_present and not calendar_sheets_present:
            missing_sheets |= set(CALENDAR_SHEETS) - available_sheets
        if missing_sheets:
            raise ImportValidationError(
                [
                    ImportIssue(
                        "file",
                        0,
                        None,
                        "missing_sheet",
                        f"Missing sheets: {', '.join(sorted(missing_sheets))}",
                    )
                ]
            )

        rows: dict[str, list[dict[str, Any]]] = {}
        sheet_issues: list[ImportIssue] = []
        selected_sheets = dict(CORE_SHEETS)
        if reference_sheets_present:
            selected_sheets.update(REFERENCE_SHEETS)
        if "Студенты" in available_sheets:
            selected_sheets.update(STUDENT_SHEETS)
        if room_sheets_present:
            selected_sheets.update(ROOM_SHEETS)
        if calendar_sheets_present:
            selected_sheets.update(CALENDAR_SHEETS)
        if constraint_sheets_present:
            selected_sheets.update(CALENDAR_CONSTRAINT_SHEETS)
        for sheet_name, section in selected_sheets.items():
            try:
                rows[section] = _read_sheet_rows(workbook[sheet_name])
            except ImportValidationError as error:
                sheet_issues.extend(error.issues)
        if sheet_issues:
            raise ImportValidationError(sheet_issues)

        return build_import_batch(
            teacher_rows=rows["teachers"],
            group_rows=rows["groups"],
            workload_rows=rows["workloads"],
            specialty_rows=rows.get("specialties"),
            curriculum_rows=rows.get("curricula"),
            discipline_rows=rows.get("disciplines"),
            student_rows=rows.get("students"),
            building_rows=rows.get("buildings"),
            room_type_rows=rows.get("room_types"),
            equipment_rows=rows.get("equipment"),
            room_rows=rows.get("rooms"),
            academic_year_rows=rows.get("academic_years"),
            calendar_period_rows=rows.get("calendar_periods"),
            bell_slot_rows=rows.get("bell_slots"),
            calendar_exception_rows=rows.get("calendar_exceptions"),
            resource_unavailability_rows=rows.get("resource_unavailability"),
        )
    finally:
        workbook.close()
